from matplotlib.colors import Colormap
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from mpl_toolkits.mplot3d import Axes3D

sFile = '0001.xlsx'
wb = load_workbook(filename = sFile , data_only = True)
ws = wb['0001']
#####################################
Fe_00950105 = []
for irow in range(1, 689):
    cell = ws.cell(row = irow, column = 1) 
    Fe_00950105.append(cell.value)    
Cr_00950105 = []
for irow in range(1, 689):
    cell = ws.cell(row = irow, column = 2) 
    Cr_00950105.append(cell.value)
Ni_00950105 = []
for irow in range(1, 689):
    cell = ws.cell(row = irow, column = 3) 
    Ni_00950105.append(cell.value)
delta_00950105 = []
for irow in range(1, 689):
    cell = ws.cell(row = irow, column = 4) 
    delta_00950105.append(cell.value)
strain_00950105 = []
for irow in range(1, 689):
    cell = ws.cell(row = irow, column = 5) 
    strain_00950105.append(cell.value)
#########################################################

Fe_00750095 = []
for irow in range(689, 2166):
    cell = ws.cell(row = irow, column = 1) 
    Fe_00750095.append(cell.value)    
Cr_00750095 = []
for irow in range(689, 2166):
    cell = ws.cell(row = irow, column = 2) 
    Cr_00750095.append(cell.value)
Ni_00750095 = []
for irow in range(689, 2166):
    cell = ws.cell(row = irow, column = 3) 
    Ni_00750095.append(cell.value)
delta_00750095 = []
for irow in range(689, 2166):
    cell = ws.cell(row = irow, column = 4) 
    delta_00750095.append(cell.value)
strain_00750095 = []
for irow in range(689, 2166):
    cell = ws.cell(row = irow, column = 5) 
    strain_00750095.append(cell.value)
##################################################
Fe_00550075 = []
for irow in range(2166, 2750):
    cell = ws.cell(row = irow, column = 1) 
    Fe_00550075.append(cell.value)    
Cr_00550075 = []
for irow in range(2166, 2750):
    cell = ws.cell(row = irow, column = 2) 
    Cr_00550075.append(cell.value)
Ni_00550075 = []
for irow in range(2166, 2750):
    cell = ws.cell(row = irow, column = 3) 
    Ni_00550075.append(cell.value)
delta_00550075 = []
for irow in range(2166, 2750):
    cell = ws.cell(row = irow, column = 4) 
    delta_00550075.append(cell.value)
strain_00550075 = []
for irow in range(2166, 2750):
    cell = ws.cell(row = irow, column = 5) 
    strain_00550075.append(cell.value)
#####################################################
Fe_01050120 = []
for irow in range(2750, 3619):
    cell = ws.cell(row = irow, column = 1) 
    Fe_01050120.append(cell.value)    
Cr_01050120 = []
for irow in range(2750, 3619):
    cell = ws.cell(row = irow, column = 2) 
    Cr_01050120.append(cell.value)
Ni_01050120 = []
for irow in range(2750, 3619):
    cell = ws.cell(row = irow, column = 3) 
    Ni_01050120.append(cell.value)
delta_01050120 = []
for irow in range(2750, 3619):
    cell = ws.cell(row = irow, column = 4) 
    delta_01050120.append(cell.value)
strain_01050120 = []
for irow in range(2750, 3619):
    cell = ws.cell(row = irow, column = 5) 
    strain_01050120.append(cell.value)
#######################################
Fe_01200150 = []
for irow in range(3619, 4763):
    cell = ws.cell(row = irow, column = 1) 
    Fe_01200150.append(cell.value)    
Cr_01200150 = []
for irow in range(3619, 4763):
    cell = ws.cell(row = irow, column = 2) 
    Cr_01200150.append(cell.value)
Ni_01200150 = []
for irow in range(3619, 4763):
    cell = ws.cell(row = irow, column = 3) 
    Ni_01200150.append(cell.value)
delta_01200150 = []
for irow in range(3619, 4763):
    cell = ws.cell(row = irow, column = 4) 
    delta_01200150.append(cell.value)
strain_01200150 = []
for irow in range(3619, 4763):
    cell = ws.cell(row = irow, column = 5) 
    strain_01200150.append(cell.value)
####################################################
Fe_00150055 = []
for irow in range(4763, 4852):
    cell = ws.cell(row = irow, column = 1) 
    Fe_00150055.append(cell.value)    
Cr_00150055 = []
for irow in range(4763, 4852):
    cell = ws.cell(row = irow, column = 2) 
    Cr_00150055.append(cell.value)
Ni_00150055 = []
for irow in range(4763, 4852):
    cell = ws.cell(row = irow, column = 3) 
    Ni_00150055.append(cell.value)
delta_00150055 = []
for irow in range(4763, 4852):
    cell = ws.cell(row = irow, column = 4) 
    delta_00150055.append(cell.value)
strain_00150055 = []
for irow in range(4763, 4852):
    cell = ws.cell(row = irow, column = 5) 
    strain_00150055.append(cell.value)
Fe_total = []
for irow in range(1, 4852):
    cell = ws.cell(row = irow, column = 1) 
    Fe_total.append(cell.value)    
Cr_total = []
for irow in range(1, 4852):
    cell = ws.cell(row = irow, column = 2) 
    Cr_total.append(cell.value)
Ni_total = []
for irow in range(1, 4852):
    cell = ws.cell(row = irow, column = 3) 
    Ni_total.append(cell.value)
delta_total = []
for irow in range(1, 4852):
    cell = ws.cell(row = irow, column = 4) 
    delta_total.append(cell.value)
strain_total = []
for irow in range(1, 4852):
    cell = ws.cell(row = irow, column = 5) 
    strain_total.append(cell.value)
######################################################
#x = Fe
#y = Cr
#z = Ni
#x_delete = Fe_delete
#y_delete = Cr_delete
#z_delete = Ni_delete

fig = plt.figure(figsize = [12,8])
ax = plt.axes(projection = "3d")
ax.set_xlim3d(0,100)
ax.set_ylim3d(0,100)
ax.set_zlim3d(0,100)
#ax.plot_surface(X, Y, Z, cmap='ocean')
ax.set_xlabel('Fe concentration(%)')
ax.set_ylabel('Cr concentration(%)')
ax.set_zlabel('Ni concentration(%)')
ax.scatter3D(Fe_00950105,Cr_00950105,Ni_00950105, color = "yellow" )
ax.scatter3D(Fe_00750095,Cr_00750095,Ni_00750095, color = "blue" )
ax.scatter3D(Fe_00550075,Cr_00550075,Ni_00550075, color = "blue")
ax.scatter3D(Fe_01050120,Cr_01050120,Ni_01050120, color = "blue")
ax.scatter3D(Fe_01200150,Cr_01200150,Ni_01200150, color = "blue")
ax.scatter3D(Fe_00150055,Cr_00150055,Ni_00150055, color = "blue")
#cmhot = plt.cm.get_cmap("rainbow")
#ax.scatter3D(Fe_total,Cr_total,Ni_total,  strain_total, cmap  = cmhot)
#print(strain_total)
fig1 = plt.figure(figsize = [12,8])
plt.xlabel('delta', fontsize = 18 )
plt.ylabel('strain', fontsize = 18)
plt.scatter(delta_00950105,strain_00950105, color = 'green')
plt.scatter(delta_00750095,strain_00750095, color = 'blue')
plt.scatter(delta_00550075,strain_00550075, color = 'yellow')
plt.scatter(delta_01050120,strain_01050120, color = 'red')
plt.scatter(delta_01200150,strain_01200150, color = 'purple')
plt.scatter(delta_00150055,strain_00150055, color = 'black')
plt.show()
